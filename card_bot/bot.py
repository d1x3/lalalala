"""
Telegram бот для безопасного хранения данных банковских карт
"""
import os
import logging
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from database import SecureCardDatabase
from ocr_parser import CardParser

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Конфигурация
BOT_TOKEN = os.getenv('BOT_TOKEN', '8080110045:AAGK01_8PByIWA-F9o4wJnlGRdQtWu89Uyo')
ALLOWED_USER_ID = os.getenv('ALLOWED_USER_ID', None)  # ID пользователя, который может использовать бота

# Инициализация БД
db = SecureCardDatabase(
    db_path='cards.db',
    key_path='.encryption_key'
)
parser = CardParser()


def check_user_access(user_id: int) -> bool:
    """Проверяет, имеет ли пользователь доступ к боту"""
    if ALLOWED_USER_ID is None:
        return True
    return str(user_id) == ALLOWED_USER_ID


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user_id = update.effective_user.id

    if not check_user_access(user_id):
        await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
        return

    welcome_text = """
🔐 **Бот для безопасного хранения банковских карт**

**Доступные команды:**
/add - Добавить карту (отправьте скриншот)
/list - Показать список карт
/excel - Экспортировать все карты в Excel
/help - Помощь

**Как использовать:**
1. Отправьте скриншот банковской карты
2. Бот автоматически распознает и сохранит данные
3. Используйте /list для просмотра карт
4. Используйте /excel для скачивания всех карт

⚠️ **Безопасность:**
- Все данные шифруются локально
- Доступ только у вас
- После отправки скриншота удалите его из чата
    """
    await update.message.reply_text(welcome_text, parse_mode='Markdown')


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /help"""
    help_text = """
📖 **Помощь**

**Команды:**
/start - Начало работы
/add - Добавить карту вручную
/list - Показать все карты
/excel - Экспортировать в Excel файл
/help - Эта справка

**Отправка скриншота:**
Просто отправьте фото карты боту. Он автоматически:
- Распознает номер карты (16 цифр)
- Найдет CVV код (3-4 цифры)
- Определит срок действия (MM/YY)

**Экспорт данных:**
- /excel - скачать красивый Excel файл с таблицей всех карт

**Советы:**
- Делайте четкие фото при хорошем освещении
- Убедитесь, что все цифры видны
- После сохранения проверьте данные
- Удаляйте экспортированные файлы после использования
    """
    await update.message.reply_text(help_text, parse_mode='Markdown')


async def list_cards(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список всех сохраненных карт"""
    user_id = update.effective_user.id
    if not check_user_access(user_id):
        await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
        return

    cards = db.get_all_cards()

    if not cards:
        await update.message.reply_text("📭 У вас пока нет сохраненных карт.")
        return

    # Создаем инлайн кнопки для каждой карты
    keyboard = []
    for card_id, card_name, created_at in cards:
        button_text = card_name if card_name else f"Карта #{card_id}"
        keyboard.append([
            InlineKeyboardButton(button_text, callback_data=f"view_{card_id}")
        ])

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "💳 **Ваши карты:**\nВыберите карту для просмотра:",
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )


async def export_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспортирует все карты в Excel файл"""
    user_id = update.effective_user.id
    if not check_user_access(user_id):
        await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
        return

    cards = db.get_all_cards()

    if not cards:
        await update.message.reply_text("📭 У вас пока нет сохраненных карт.")
        return

    processing_msg = await update.message.reply_text("📊 Создаю Excel файл...")

    try:
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Мои карты"

        # Заголовки
        headers = ["ID", "Название", "Номер карты", "CVV", "Срок действия", "Дата добавления"]
        ws.append(headers)

        # Стилизация заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Добавляем данные карт
        for card_id, card_name, created_at in cards:
            card_data = db.get_card(card_id)
            if card_data:
                ws.append([
                    card_id,
                    card_name if card_name else f"Карта #{card_id}",
                    card_data['card_number'],
                    card_data['cvv'],
                    card_data['expiry'],
                    created_at
                ])

        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20

        # Выравнивание данных
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Сохраняем файл
        filename = f"my_cards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

        # Отправляем файл пользователю
        with open(filename, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename="my_cards.xlsx",
                caption="📊 **Ваши карты в Excel файле**\n\n"
                       "⚠️ **ВАЖНО:**\n"
                       "- Удалите файл после использования!\n"
                       "- Не храните его в открытом виде\n"
                       "- Не отправляйте никому"
            )

        # Удаляем временный файл
        os.remove(filename)

        await processing_msg.delete()

    except Exception as e:
        logger.error(f"Ошибка при создании Excel: {e}")
        await processing_msg.edit_text(
            f"❌ Ошибка при создании Excel файла:\n`{str(e)}`",
            parse_mode='Markdown'
        )


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает отправленное фото карты"""
    user_id = update.effective_user.id
    if not check_user_access(user_id):
        await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
        return

    # Отправляем сообщение о начале обработки
    processing_msg = await update.message.reply_text("🔄 Обрабатываю изображение...")

    try:
        # Получаем фото (берем самое большое разрешение)
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)

        # Скачиваем фото
        image_bytes = await file.download_as_bytearray()

        # Парсим ВСЕ карты с изображения
        all_cards = parser.parse_all_cards_from_image(bytes(image_bytes))

        # Проверяем, что удалось распознать хотя бы одну карту
        if not all_cards:
            await processing_msg.edit_text(
                "❌ Не удалось распознать номер карты.\n\n"
                "💡 Попробуйте:\n"
                "- Сделать более четкое фото\n"
                "- Убедиться, что все цифры видны\n"
                "- Улучшить освещение\n"
                "- Отправить фото с белым фоном"
            )
            return

        # Если найдено больше одной карты
        if len(all_cards) > 1:
            await processing_msg.edit_text(f"🎉 Найдено карт: {len(all_cards)}\nОбрабатываю каждую...")

        # Обрабатываем каждую найденную карту
        saved_count = 0
        duplicate_count = 0

        for idx, card_data in enumerate(all_cards, 1):
            card_number = card_data.get('card_number')
            cvv = card_data.get('cvv')
            expiry = card_data.get('expiry')

            if not card_number:
                continue

            # Проверка на дубликат
            if db.card_exists(card_number):
                duplicate_count += 1
                result_text = f"⚠️ **Карта {idx}/{len(all_cards)}**\n\n"
                result_text += f"💳 Номер: `{card_number}`\n\n"
                result_text += "❌ **Эта карта уже есть в базе данных!**"
                await update.message.reply_text(result_text, parse_mode='Markdown')
                continue

            # Валидация номера карты по алгоритму Луна
            is_valid = parser.validate_card_number(card_number)

            # Показываем распознанные данные
            result_text = f"✅ **Карта {idx}/{len(all_cards)} распознана:**\n\n"
            result_text += f"💳 Номер карты: `{card_number}`\n"
            result_text += f"🔐 CVV: `{cvv or 'Не найден'}`\n"
            result_text += f"📅 Срок: `{expiry or 'Не найден'}`\n\n"

            if not is_valid:
                result_text += "⚠️ Номер не прошел проверку по алгоритму Луна\n"

            # Если не все данные распознаны
            if not cvv or not expiry:
                result_text += "⚠️ Не все данные распознаны\n\n"

            result_text += "⚠️ **Проверьте CVV код!** OCR может ошибаться.\n\n"
            result_text += "Если CVV неправильный:\n"
            result_text += "1. Нажмите 'Пропустить'\n"
            result_text += "2. Введите данные вручную:\n"
            result_text += f"`{card_number} ВАШ_CVV {expiry or 'MM/YY'}`\n\n"
            result_text += "Сохранить эту карту?"

            # Кнопки для сохранения
            keyboard = [
                [
                    InlineKeyboardButton("✅ Сохранить", callback_data=f"save_{card_number}_{cvv}_{expiry}"),
                    InlineKeyboardButton("❌ Пропустить", callback_data="cancel")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)

            await update.message.reply_text(
                result_text,
                reply_markup=reply_markup,
                parse_mode='Markdown'
            )

        # Удаляем сообщение об обработке
        await processing_msg.delete()

        # Итоговое сообщение
        if len(all_cards) > 1 or duplicate_count > 0:
            summary = f"📊 **Итого:**\n"
            summary += f"Найдено карт: {len(all_cards)}\n"
            if duplicate_count > 0:
                summary += f"Дубликатов: {duplicate_count}\n"
            await update.message.reply_text(summary, parse_mode='Markdown')

    except Exception as e:
        logger.error(f"Ошибка обработки фото: {e}")
        await processing_msg.edit_text(
            f"❌ Произошла ошибка при обработке изображения:\n`{str(e)}`",
            parse_mode='Markdown'
        )


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает нажатия на инлайн кнопки"""
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    if not check_user_access(user_id):
        await query.edit_message_text("⛔ У вас нет доступа к этому боту.")
        return

    data = query.data

    # Отмена
    if data == "cancel":
        await query.edit_message_text("❌ Операция отменена.")
        return

    # Сохранение карты
    if data.startswith("save_"):
        parts = data.split("_")
        if len(parts) >= 4:
            card_number = parts[1]
            cvv = parts[2] if parts[2] != 'None' else None
            expiry = parts[3] if parts[3] != 'None' else None

            # Проверка на дубликат
            if db.card_exists(card_number):
                await query.edit_message_text(
                    f"⚠️ **Карта уже существует!**\n\n"
                    f"💳 Номер: `{card_number}`\n\n"
                    f"❌ Эта карта уже есть в базе данных.\n"
                    f"Используйте /list для просмотра.",
                    parse_mode='Markdown'
                )
                return

            # Сохраняем в БД
            card_id = db.add_card(
                card_number=card_number,
                cvv=cvv,
                expiry=expiry,
                card_name=f"Карта {card_number[-4:]}"
            )

            await query.edit_message_text(
                f"✅ Карта сохранена!\n\n"
                f"ID: {card_id}\n"
                f"Используйте /list для просмотра всех карт."
            )
            return

    # Просмотр карты
    if data.startswith("view_"):
        card_id = int(data.split("_")[1])
        card_data = db.get_card(card_id)

        if not card_data:
            await query.edit_message_text("❌ Карта не найдена.")
            return

        card_text = f"💳 **Карта #{card_id}**\n\n"
        card_text += f"**Номер:** `{card_data['card_number']}`\n"
        card_text += f"**CVV:** `{card_data['cvv']}`\n"
        card_text += f"**Срок:** `{card_data['expiry']}`\n\n"
        card_text += "Нажмите на данные для копирования"

        # Кнопки для управления
        keyboard = [
            [InlineKeyboardButton("🗑 Удалить", callback_data=f"delete_{card_id}")],
            [InlineKeyboardButton("◀️ Назад", callback_data="back_to_list")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(
            card_text,
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return

    # Удаление карты
    if data.startswith("delete_"):
        card_id = int(data.split("_")[1])

        keyboard = [
            [
                InlineKeyboardButton("✅ Да, удалить", callback_data=f"confirm_delete_{card_id}"),
                InlineKeyboardButton("❌ Отмена", callback_data=f"view_{card_id}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(
            "⚠️ Вы уверены, что хотите удалить эту карту?",
            reply_markup=reply_markup
        )
        return

    # Подтверждение удаления
    if data.startswith("confirm_delete_"):
        card_id = int(data.split("_")[2])
        if db.delete_card(card_id):
            await query.edit_message_text("✅ Карта удалена.")
        else:
            await query.edit_message_text("❌ Ошибка при удалении карты.")
        return

    # Назад к списку
    if data == "back_to_list":
        cards = db.get_all_cards()

        if not cards:
            await query.edit_message_text("📭 У вас нет сохраненных карт.")
            return

        keyboard = []
        for card_id, card_name, created_at in cards:
            button_text = card_name if card_name else f"Карта #{card_id}"
            keyboard.append([
                InlineKeyboardButton(button_text, callback_data=f"view_{card_id}")
            ])

        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            "💳 **Ваши карты:**\nВыберите карту для просмотра:",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return


async def add_card_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавление карты вручную"""
    user_id = update.effective_user.id
    if not check_user_access(user_id):
        await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
        return

    help_text = """
📝 **Добавление карты вручную**

Отправьте данные в формате:
`номер CVV MM/YY`

Пример:
`4276 3801 2345 6789 123 12/25`

Или просто отправьте скриншот карты для автоматического распознавания.
    """
    await update.message.reply_text(help_text, parse_mode='Markdown')


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка текстовых сообщений для ручного добавления карт"""
    user_id = update.effective_user.id
    if not check_user_access(user_id):
        return

    text = update.message.text

    # Парсим данные из текста
    card_number = parser.parse_card_number(text)
    cvv = parser.parse_cvv(text, card_number)
    expiry = parser.parse_expiry(text)

    if card_number:
        # Проверка на дубликат
        if db.card_exists(card_number):
            await update.message.reply_text(
                f"⚠️ **Карта уже существует!**\n\n"
                f"💳 Номер: `{card_number}`\n\n"
                f"❌ Эта карта уже есть в базе данных.\n"
                f"Используйте /list для просмотра.",
                parse_mode='Markdown'
            )
            return

        # Сохраняем карту
        card_id = db.add_card(
            card_number=card_number,
            cvv=cvv,
            expiry=expiry,
            card_name=f"Карта {card_number[-4:]}"
        )

        result_text = f"✅ Карта сохранена!\n\n"
        result_text += f"ID: {card_id}\n"
        result_text += f"💳 Номер: `{card_number}`\n"
        result_text += f"🔐 CVV: `{cvv or 'Не указан'}`\n"
        result_text += f"📅 Срок: `{expiry or 'Не указан'}`"

        await update.message.reply_text(result_text, parse_mode='Markdown')
    else:
        await update.message.reply_text(
            "❌ Не удалось распознать данные карты.\n"
            "Используйте формат: `номер CVV MM/YY`",
            parse_mode='Markdown'
        )


def main():
    """Запуск бота"""
    # Создаем приложение
    application = Application.builder().token(BOT_TOKEN).build()

    # Регистрируем обработчики команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("list", list_cards))
    application.add_handler(CommandHandler("excel", export_to_excel))
    application.add_handler(CommandHandler("add", add_card_manual))

    # Обработчики сообщений
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    # Обработчик кнопок
    application.add_handler(CallbackQueryHandler(button_callback))

    # Запускаем бота
    logger.info("🤖 Бот запущен!")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
